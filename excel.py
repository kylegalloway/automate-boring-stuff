import openpyxl
import os

os.chdir('/mnt/c/Users/ckgallow/Downloads')

# workbook = openpyxl.load_workbook('example.xlsx')

# print(workbook.get_sheet_names())
# sheet = workbook.get_sheet_by_name('Sheet1')

# print(sheet['B1'].value)
# print(sheet.cell(row=1, column=2).value)

# for i in range(1, 8):
#     print(i, sheet.cell(row=i, column=2).value)


wb = openpyxl.Workbook()
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet')
sheet2 = wb.create_sheet()
print(wb.get_sheet_names())
sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names())
sheet2 = wb.create_sheet(index=0, title='My Other Sheet')
print(wb.get_sheet_names())
sheet['A1'] = 42
sheet['A2'] = 'Hello'
sheet['B2'] = 'World'

wb.save('example-written.xlsx')
